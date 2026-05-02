#!/usr/bin/env python3
"""
Bank Statement Analyzer — Local Tool
Runs entirely on your machine. No internet, no AI model required.
Reads your CSVs + category_keywords.csv → produces Excel dashboard.

Usage:
    python bank_analyzer.py                  # Interactive mode (guided prompts)
    python bank_analyzer.py --config my.json # Use saved config

Requirements:
    pip install pandas openpyxl
"""

import os, sys, json, re, argparse
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
KEYWORDS_CSV = SCRIPT_DIR / "category_keywords.csv"
CONFIGS_FILE = SCRIPT_DIR / "statement_configs.json"
SESSION_FILE = SCRIPT_DIR / "session.json"


# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION  —  remembers which files were added and with what config
# ═══════════════════════════════════════════════════════════════════════════════

def load_session():
    """Return session dict with list of previously added files."""
    if SESSION_FILE.exists():
        try:
            with open(SESSION_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"files": [], "output": ""}

def save_session(file_cfgs, out_path):
    """Persist the current set of (filepath, cfg) pairs."""
    session = {
        "files": [
            {
                "path"       : str(Path(fp).resolve()),
                "source_name": cfg.get("name", ""),
                "cfg"        : cfg,
            }
            for fp, cfg in file_cfgs
        ],
        "output"  : str(out_path),
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "keywords_mtime": (
            os.path.getmtime(KEYWORDS_CSV) if KEYWORDS_CSV.exists() else None
        ),
    }
    with open(SESSION_FILE, "w", encoding="utf-8") as f:
        json.dump(session, f, indent=2)
    print(f"  Session saved  -> {SESSION_FILE}")

def session_file_key(fp):
    """Canonical key for de-duplication: resolved absolute path string."""
    return str(Path(fp).resolve())

# ─── Colours (for Excel) ──────────────────────────────────────────────────────
CLR = {
    "header_bg"   : "1F2937",
    "header_fg"   : "FFFFFF",
    "subheader_bg": "374151",
    "subheader_fg": "F9FAFB",
    "alt_row"     : "F8FAFC",
    "debit"       : "DC2626",
    "credit"      : "16A34A",
    "neutral"     : "374151",
    "accent"      : "2563EB",
    "border"      : "E2E8F0",
    "warn_bg"     : "FEF3C7",
    "warn_fg"     : "92400E",
}

CAT_COLORS = {
    "Food & Dining"   : "378ADD", "Shopping"       : "1D9E75",
    "Transport"       : "D85A30", "Entertainment"  : "7F77DD",
    "Utilities"       : "BA7517", "Health"         : "E24B4A",
    "Travel"          : "0F6E56", "Salary/Income"  : "639922",
    "EMI/Loan"        : "D4537E", "Transfers"      : "888780",
    "Investment"      : "185FA5", "Rent"           : "9B59B6",
    "Education"       : "E67E22", "Subscriptions"  : "16A085",
    "Kids & Family"   : "F39C12", "Personal Care"  : "E91E63",
    "ATM & Cash"      : "607D8B", "Charity"        : "795548",
    "Other"           : "B4B2A9",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  1. CATEGORY ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def load_keywords(path=KEYWORDS_CSV):
    """Load category → [keyword, ...] mapping from CSV."""
    if not path.exists():
        print(f"[WARN] {path} not found — using built-in defaults.")
        return {}
    df = pd.read_csv(path, dtype=str).fillna("")
    mapping = {}
    for _, row in df.iterrows():
        cat  = row["Category"].strip()
        kws  = [k.strip().lower() for k in row["Keywords"].split(",") if k.strip()]
        mapping[cat] = kws
    return mapping


def categorize(description: str, keyword_map: dict) -> str:
    d = str(description).lower()
    for category, keywords in keyword_map.items():
        if category in ("Other", ""):
            continue
        for kw in keywords:
            if kw and kw in d:
                return category
    return "Other"


# ═══════════════════════════════════════════════════════════════════════════════
#  2. CONFIG — statement column layouts
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_CONFIGS = {
    "HDFC Bank": {
        "type": "bank", "skip_rows": 0, "date_col": "Date",
        "date_format": "%d/%m/%y", "desc_col": "Narration",
        "amount_format": "separate",
        "debit_col": "Withdrawal Amt.", "credit_col": "Deposit Amt.",
    },
    "ICICI Bank": {
        "type": "bank", "skip_rows": 0, "date_col": "Transaction Date",
        "date_format": "%d/%m/%Y", "desc_col": "Remarks",
        "amount_format": "separate",
        "debit_col": "Withdrawal Amount (INR )", "credit_col": "Deposit Amount (INR )",
    },
    "SBI Bank": {
        "type": "bank", "skip_rows": 0, "date_col": "Txn Date",
        "date_format": "%d %b %Y", "desc_col": "Description",
        "amount_format": "separate",
        "debit_col": "Debit", "credit_col": "Credit",
    },
    "Axis Bank": {
        "type": "bank", "skip_rows": 0, "date_col": "Tran Date",
        "date_format": "%d-%m-%Y", "desc_col": "PARTICULARS",
        "amount_format": "separate",
        "debit_col": "DR", "credit_col": "CR",
    },
    "HDFC Credit Card": {
        "type": "credit", "skip_rows": 0, "date_col": "Date",
        "date_format": "%d/%m/%Y", "desc_col": "Description",
        "amount_format": "signed",
        "amount_col": "Amount", "sign_convention": "pos_is_debit",
    },
    "ICICI Credit Card": {
        "type": "credit", "skip_rows": 0, "date_col": "Date",
        "date_format": "%d/%m/%Y", "desc_col": "Details",
        "amount_format": "signed",
        "amount_col": "Amount (in Rs.)", "sign_convention": "pos_is_debit",
    },
    "Kotak Bank": {
        "type": "bank", "skip_rows": 0, "date_col": "Date",
        "date_format": "%d-%m-%Y", "desc_col": "Description",
        "amount_format": "separate",
        "debit_col": "Debit", "credit_col": "Credit",
    },
    "Custom": {
        "type": "bank", "skip_rows": 0, "date_col": "",
        "date_format": "%d/%m/%Y", "desc_col": "",
        "amount_format": "separate",
        "debit_col": "", "credit_col": "",
    },
}

def load_configs():
    if CONFIGS_FILE.exists():
        with open(CONFIGS_FILE) as f:
            saved = json.load(f)
        merged = {**DEFAULT_CONFIGS, **saved}
        return merged
    return DEFAULT_CONFIGS.copy()

def save_configs(configs):
    # Only save user-added/modified ones
    user_configs = {k: v for k, v in configs.items() if k not in DEFAULT_CONFIGS
                    or v != DEFAULT_CONFIGS.get(k)}
    with open(CONFIGS_FILE, "w") as f:
        json.dump(user_configs, f, indent=2)


# ═══════════════════════════════════════════════════════════════════════════════
#  3. CSV PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def clean_amount(val):
    """Convert '1,23,456.78' / '(500)' / '-500' / '' → float."""
    if pd.isna(val) or str(val).strip() in ("", "-", "--"):
        return 0.0
    s = str(val).replace(",", "").replace("₹", "").replace("Rs.", "").strip()
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        v = float(s)
        return -v if negative else v
    except ValueError:
        return 0.0


def find_col(df, name):
    """Find column by name substring (case-insensitive) or integer index (1-based)."""
    if not name:
        return None
    # Try integer index
    try:
        idx = int(name) - 1
        if 0 <= idx < len(df.columns):
            return df.columns[idx]
    except (ValueError, TypeError):
        pass
    # Try exact match
    if name in df.columns:
        return name
    # Try case-insensitive substring
    name_l = name.lower().strip()
    for col in df.columns:
        if name_l in col.lower().strip():
            return col
    return None



# All date formats to try when auto-detecting
DATE_FORMATS = [
    "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
    "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
    "%d %b %y", "%d/%b/%Y", "%d-%b-%Y", "%d-%b-%y",
    "%Y/%m/%d",
]

def try_parse_date(raw: str):
    """Try all known date formats; return (datetime, format_used) or (None, None)."""
    raw = raw.strip()
    # Strip time portion if present: "25/03/2024 14:30:00" → "25/03/2024"
    raw_date_only = raw.split(" ")[0] if " " in raw else raw
    raw_date_only = raw_date_only.split("T")[0]  # ISO 8601
    for fmt in DATE_FORMATS:
        for attempt in (raw, raw_date_only):
            try:
                return datetime.strptime(attempt, fmt), fmt
            except Exception:
                continue
    return None, None


def auto_detect_columns(df, cfg):
    """
    For each config key, try to find a matching column.
    Also try common fallback names if the config name doesn't match.
    Returns dict of resolved column names.
    """
    DATE_ALIASES  = ["date", "transaction date", "txn date", "value date",
                     "posting date", "tran date", "trans date", "dated"]
    DESC_ALIASES  = ["description", "narration", "details", "particulars",
                     "remarks", "transaction details", "transaction description",
                     "merchant", "transaction narration", "trans details",
                     "transaction remarks", "info"]
    DEBIT_ALIASES = ["debit", "withdrawal", "dr", "debit amount", "withdrawal amt",
                     "withdrawal amount", "debit amt", "amount debited",
                     "spent", "charges"]
    CREDIT_ALIASES= ["credit", "deposit", "cr", "credit amount", "deposit amt",
                     "deposit amount", "credit amt", "amount credited", "payments"]
    AMT_ALIASES   = ["amount (in rs.)", "amount (inr)", "inr amount", "amount (rs)",
                     "transaction amount", "txn amount", "net amount",
                     "amount", "amt"]

    def resolve(cfg_key, aliases):
        # 1. Try what's in the config first
        col = find_col(df, cfg.get(cfg_key, ""))
        if col:
            return col
        # 2. Try aliases
        for alias in aliases:
            col = find_col(df, alias)
            if col:
                return col
        return None

    return {
        "date_col"  : resolve("date_col",   DATE_ALIASES),
        "desc_col"  : resolve("desc_col",   DESC_ALIASES),
        "debit_col" : resolve("debit_col",  DEBIT_ALIASES),
        "credit_col": resolve("credit_col", CREDIT_ALIASES),
        "amount_col": resolve("amount_col", AMT_ALIASES),
    }


def parse_statement(filepath, cfg, keyword_map):
    """Parse one CSV file according to cfg; return list of dicts."""
    path = Path(filepath)
    skip = int(cfg.get("skip_rows", 0))

    # ── Load CSV ──────────────────────────────────────────────────────────────
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(path, skiprows=skip, dtype=str, encoding=enc,
                             on_bad_lines="skip")
            break
        except UnicodeDecodeError:
            continue
    else:
        print(f"  [ERROR] Could not read {path.name} with any encoding.")
        return []

    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    if df.empty:
        print(f"  [WARN] {path.name} is empty after loading.")
        return []

    # ── Auto-detect columns ───────────────────────────────────────────────────
    print(f"\n  Columns found in {path.name}:")
    print(f"    {list(df.columns)}")

    resolved = auto_detect_columns(df, cfg)
    date_col  = resolved["date_col"]
    desc_col  = resolved["desc_col"]

    print(f"  Mapped → date='{date_col}'  desc='{desc_col}'", end="")

    afmt = cfg.get("amount_format", "separate")
    if afmt == "separate":
        dc = resolved["debit_col"]
        cc = resolved["credit_col"]
        print(f"  debit='{dc}'  credit='{cc}'")
    else:
        ac = resolved["amount_col"]
        print(f"  amount='{ac}'  sign={cfg.get('sign_convention','pos_is_debit')}")

    if not date_col:
        print(f"  [ERROR] Cannot find a date column in {path.name}.")
        print(f"          Please run again, choose NEW, and specify the exact column name.")
        return []

    # ── Auto-detect date format from first valid date value ───────────────────
    configured_fmt = cfg.get("date_format", "")
    detected_fmt   = None
    sample_date    = None

    for val in df[date_col].dropna():
        raw = str(val).strip()
        if not raw or raw.lower() in ("nan", "date", ""):
            continue
        sample_date = raw
        # Try configured format first
        if configured_fmt:
            try:
                datetime.strptime(raw.split(" ")[0].split("T")[0], configured_fmt)
                detected_fmt = configured_fmt
                break
            except Exception:
                pass
        # Fall back to auto-detection
        _, fmt_found = try_parse_date(raw)
        if fmt_found:
            detected_fmt = fmt_found
            break

    if detected_fmt:
        print(f"  Date format detected: '{detected_fmt}'  (sample: '{sample_date}')")
    else:
        print(f"  [WARN] Could not detect date format from sample '{sample_date}'.")
        print(f"         These rows will be tagged month='Unknown'.")

    # ── Parse rows ────────────────────────────────────────────────────────────
    rows = []
    unknown_dates = 0

    for _, row in df.iterrows():
        # Date
        raw_date = str(row[date_col]).strip() if date_col else ""
        date_obj = None
        date_str = raw_date
        month    = "Unknown"

        if raw_date and raw_date.lower() not in ("nan", ""):
            if detected_fmt:
                raw_clean = raw_date.split(" ")[0].split("T")[0]
                try:
                    date_obj = datetime.strptime(raw_clean, detected_fmt)
                except Exception:
                    date_obj, _ = try_parse_date(raw_date)
            else:
                date_obj, _ = try_parse_date(raw_date)

            if date_obj:
                date_str = date_obj.strftime("%Y-%m-%d")
                month    = date_obj.strftime("%b %Y")
            else:
                unknown_dates += 1

        # Description
        desc = str(row[desc_col]).strip() if desc_col and desc_col in row else ""

        # Amount
        debit = credit = 0.0
        if afmt == "separate":
            debit  = clean_amount(row[dc]) if dc and dc in row else 0.0
            credit = clean_amount(row[cc]) if cc and cc in row else 0.0
            debit, credit = abs(debit), abs(credit)
        else:
            amt  = clean_amount(row[ac]) if ac and ac in row else 0.0
            sign = cfg.get("sign_convention", "pos_is_debit")
            if sign == "pos_is_debit":
                debit  = amt if amt > 0 else 0.0
                credit = abs(amt) if amt < 0 else 0.0
            else:
                credit = amt if amt > 0 else 0.0
                debit  = abs(amt) if amt < 0 else 0.0

        if debit == 0 and credit == 0:
            continue

        amount   = debit if debit > 0 else credit
        is_debit = debit > 0

        rows.append({
            "date"       : date_str,
            "date_obj"   : date_obj,
            "month"      : month,
            "description": desc,
            "amount"     : round(amount, 2),
            "is_debit"   : is_debit,
            "category"   : categorize(desc, keyword_map),
            "source"     : path.name,
            "source_name": cfg.get("name", path.stem),
            "type"       : cfg.get("type", "bank"),
        })

    if unknown_dates:
        print(f"  [WARN] {unknown_dates} rows had unrecognised dates → tagged 'Unknown'.")

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
#  4. EXCEL DASHBOARD BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def _border(sides="all", color="D1D5DB"):
    s = Side(style="thin", color=color)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    if sides == "top_bottom":
        return Border(top=s, bottom=s)
    return Border()

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def header_row(ws, row_num, cols, bg=CLR["header_bg"], fg=CLR["header_fg"]):
    for ci, text in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=ci, value=text)
        c.font      = _font(bold=True, color=fg, size=11)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()

def data_cell(ws, row, col, value, fmt=None, bold=False,
              color="1F2937", align="left", bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=color, size=10)
    c.alignment = _align(align)
    c.border    = _border("bottom", CLR["border"])
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = _fill(bg)
    return c


# ── Sheet helpers ──────────────────────────────────────────────────────────────

def build_summary_sheet(wb, df):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    debits  = df[df.is_debit]
    credits = df[~df.is_debit]
    total_spend  = debits.amount.sum()
    total_income = credits.amount.sum()
    net_savings  = total_income - total_spend
    months = df.month.nunique()
    avg_monthly_spend  = total_spend  / months if months else 0
    avg_monthly_income = total_income / months if months else 0

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value     = "Personal Finance Dashboard"
    title.font      = _font(bold=True, size=20, color=CLR["header_fg"])
    title.fill      = _fill(CLR["header_bg"])
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value     = (f"Generated on {datetime.now():%d %B %Y}   |   "
                     f"{len(df):,} transactions   |   {months} month(s) analysed")
    sub.font      = _font(size=10, color="9CA3AF", italic=True)
    sub.fill      = _fill(CLR["subheader_bg"])
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 12  # spacer

    # ── Section label: SPEND ───────────────────────────────────────────────────
    ws.merge_cells("A4:E4")
    sl = ws["A4"]
    sl.value = "  SPEND SUMMARY"
    sl.font  = _font(bold=True, size=10, color="FFFFFF")
    sl.fill  = _fill("991B1B")   # dark red
    sl.alignment = _align("left")
    ws.row_dimensions[4].height = 20

    # ── KPI row — spend (row 5 label, row 6 value) ────────────────────────────
    spend_kpis = [
        ("Total Spend",          total_spend,         CLR["debit"],   "₹#,##0"),
        ("Avg Monthly Spend",    avg_monthly_spend,   "C2410C",       "₹#,##0"),
        ("Spend Transactions",   len(debits),         "7F1D1D",       "0"),
        ("Top Spend Category",   debits.groupby("category")["amount"].sum().idxmax()
                                 if not debits.empty else "—",        "991B1B", None),
        ("Savings Rate",         (net_savings/total_income*100) if total_income else 0,
                                                      "15803D",       "0.0%"),
    ]

    for ci, (label, value, color, fmt) in enumerate(spend_kpis, 1):
        lc = ws.cell(row=5, column=ci, value=label)
        lc.font      = _font(size=9, color="6B7280")
        lc.alignment = _align("center")
        lc.fill      = _fill("FEF2F2")
        lc.border    = _border("bottom", "FECACA")

        vc = ws.cell(row=6, column=ci, value=value)
        vc.font      = _font(bold=True, size=13, color=color)
        vc.alignment = _align("center")
        vc.fill      = _fill("FEF2F2")
        vc.border    = _border("bottom", color)
        if fmt and isinstance(value, (int, float)):
            vc.number_format = fmt

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 10  # spacer

    # ── Section label: INCOME ─────────────────────────────────────────────────
    ws.merge_cells("A8:E8")
    il = ws["A8"]
    il.value = "  INCOME SUMMARY"
    il.font  = _font(bold=True, size=10, color="FFFFFF")
    il.fill  = _fill("14532D")   # dark green
    il.alignment = _align("left")
    ws.row_dimensions[8].height = 20

    income_cat_totals = (credits.groupby("category")["amount"].sum()
                         if not credits.empty else pd.Series(dtype=float))
    top_income_cat = income_cat_totals.idxmax() if not income_cat_totals.empty else "—"

    income_kpis = [
        ("Total Income",         total_income,        CLR["credit"],  "₹#,##0"),
        ("Avg Monthly Income",   avg_monthly_income,  "166534",       "₹#,##0"),
        ("Income Transactions",  len(credits),        "14532D",       "0"),
        ("Top Income Source",    top_income_cat,      "166534",       None),
        ("Net Savings",          net_savings,
                                 CLR["credit"] if net_savings >= 0 else CLR["debit"],
                                                                      "₹#,##0"),
    ]

    for ci, (label, value, color, fmt) in enumerate(income_kpis, 1):
        lc = ws.cell(row=9, column=ci, value=label)
        lc.font      = _font(size=9, color="6B7280")
        lc.alignment = _align("center")
        lc.fill      = _fill("F0FDF4")
        lc.border    = _border("bottom", "BBF7D0")

        vc = ws.cell(row=10, column=ci, value=value)
        vc.font      = _font(bold=True, size=13, color=color)
        vc.alignment = _align("center")
        vc.fill      = _fill("F0FDF4")
        vc.border    = _border("bottom", color)
        if fmt and isinstance(value, (int, float)):
            vc.number_format = fmt

    ws.row_dimensions[9].height  = 18
    ws.row_dimensions[10].height = 30
    ws.row_dimensions[11].height = 14  # spacer

    # ── SPEND by category table ────────────────────────────────────────────────
    ws.merge_cells("A12:E12")
    sh = ws["A12"]
    sh.value = "  SPEND BY CATEGORY"
    sh.font  = _font(bold=True, size=10, color="FFFFFF")
    sh.fill  = _fill(CLR["header_bg"])
    sh.alignment = _align("left")
    ws.row_dimensions[12].height = 20

    cat_totals = (debits.groupby("category")["amount"]
                  .agg(["sum", "count"])
                  .rename(columns={"sum": "total", "count": "txn_count"})
                  .sort_values("total", ascending=False))
    cat_totals["pct"] = cat_totals["total"] / total_spend if total_spend else 0

    header_row(ws, 13, ["Category", "Transactions", "Total Spend", "% of Spend",
                         "Avg per Transaction"], bg="374151", fg="F9FAFB")
    ws.row_dimensions[13].height = 20

    for ri, (cat, row) in enumerate(cat_totals.iterrows(), 14):
        bg = "FFFFFF" if ri % 2 == 0 else "FFF5F5"
        hex_c = CAT_COLORS.get(cat, "B4B2A9")
        # Colour swatch in column A via cell background on a narrow col
        swatch = ws.cell(row=ri, column=1, value=f"  {cat}")
        swatch.font      = _font(bold=True, size=10, color=hex_c)
        swatch.fill      = _fill(bg)
        swatch.border    = _border("bottom", CLR["border"])
        swatch.alignment = _align("left")
        data_cell(ws, ri, 2, int(row.txn_count), align="center", bg=bg)
        data_cell(ws, ri, 3, row.total, fmt="₹#,##0", align="right", bg=bg)
        data_cell(ws, ri, 4, row.pct, fmt="0.0%", align="center", bg=bg)
        avg = row.total / row.txn_count if row.txn_count else 0
        data_cell(ws, ri, 5, avg, fmt="₹#,##0", align="right", bg=bg)

    spend_last = 13 + len(cat_totals) + 1
    ws.row_dimensions[spend_last].height = 14  # spacer

    # ── INCOME by category table ───────────────────────────────────────────────
    inc_start = spend_last + 1
    ws.merge_cells(f"A{inc_start}:E{inc_start}")
    ih = ws[f"A{inc_start}"]
    ih.value = "  INCOME BY CATEGORY"
    ih.font  = _font(bold=True, size=10, color="FFFFFF")
    ih.fill  = _fill("166534")
    ih.alignment = _align("left")
    ws.row_dimensions[inc_start].height = 20

    inc_totals = (credits.groupby("category")["amount"]
                  .agg(["sum", "count"])
                  .rename(columns={"sum": "total", "count": "txn_count"})
                  .sort_values("total", ascending=False))
    inc_totals["pct"] = inc_totals["total"] / total_income if total_income else 0

    header_row(ws, inc_start + 1, ["Category", "Transactions", "Total Income",
                                    "% of Income", "Avg per Transaction"],
               bg="166534", fg="F9FAFB")
    ws.row_dimensions[inc_start + 1].height = 20

    if inc_totals.empty:
        ws.merge_cells(f"A{inc_start+2}:E{inc_start+2}")
        nc = ws[f"A{inc_start+2}"]
        nc.value = "No income transactions found"
        nc.font  = _font(size=10, color="6B7280", italic=True)
        nc.alignment = _align("center")
    else:
        for ri, (cat, row) in enumerate(inc_totals.iterrows(), inc_start + 2):
            bg = "FFFFFF" if ri % 2 == 0 else "F0FDF4"
            hex_c = CAT_COLORS.get(cat, "B4B2A9")
            swatch = ws.cell(row=ri, column=1, value=f"  {cat}")
            swatch.font      = _font(bold=True, size=10, color=hex_c)
            swatch.fill      = _fill(bg)
            swatch.border    = _border("bottom", CLR["border"])
            swatch.alignment = _align("left")
            data_cell(ws, ri, 2, int(row.txn_count), align="center", bg=bg)
            data_cell(ws, ri, 3, row.total, fmt="₹#,##0", align="right", bg=bg)
            data_cell(ws, ri, 4, row.pct, fmt="0.0%", align="center", bg=bg)
            avg = row.total / row.txn_count if row.txn_count else 0
            data_cell(ws, ri, 5, avg, fmt="₹#,##0", align="right", bg=bg)

    set_col_widths(ws, {"A": 24, "B": 16, "C": 18, "D": 14, "E": 22,
                        "F": 2, "G": 2, "H": 2, "I": 2})
    ws.freeze_panes = "A14"
    return ws


def build_monthly_sheet(wb, df):
    ws = wb.create_sheet("Monthly Breakdown")
    ws.sheet_view.showGridLines = False

    debits  = df[df.is_debit]
    credits = df[~df.is_debit]

    # Month order
    def month_sort(m):
        try: return datetime.strptime(m, "%b %Y")
        except: return datetime.min

    months_sorted = sorted(df.month.unique(), key=month_sort)
    categories    = sorted(debits.category.unique())

    # Build pivot: rows=month, cols=category
    pivot = (debits.groupby(["month", "category"])["amount"]
             .sum().unstack(fill_value=0))
    pivot = pivot.reindex(months_sorted).fillna(0)

    header_row(ws, 1, ["Month", "Total Spend", "Total Income", "Net", "# Txns"]
               + list(pivot.columns))
    ws.row_dimensions[1].height = 22

    month_debits  = debits.groupby("month")["amount"].sum()
    month_credits = credits.groupby("month")["amount"].sum()
    month_txns    = df.groupby("month").size()

    for ri, month in enumerate(months_sorted, 2):
        spend  = month_debits.get(month, 0)
        income = month_credits.get(month, 0)
        net    = income - spend
        txns   = month_txns.get(month, 0)
        bg = "FFFFFF" if ri % 2 == 0 else CLR["alt_row"]

        data_cell(ws, ri, 1, month, bold=True, bg=bg)
        data_cell(ws, ri, 2, spend, fmt="₹#,##0", align="right",
                  color=CLR["debit"], bg=bg)
        data_cell(ws, ri, 3, income, fmt="₹#,##0", align="right",
                  color=CLR["credit"], bg=bg)
        net_color = CLR["credit"] if net >= 0 else CLR["debit"]
        data_cell(ws, ri, 4, net, fmt="₹#,##0", align="right",
                  color=net_color, bg=bg, bold=True)
        data_cell(ws, ri, 5, int(txns), align="center", bg=bg)

        for ci, cat in enumerate(pivot.columns, 6):
            val = pivot.loc[month, cat] if month in pivot.index else 0
            data_cell(ws, ri, ci, val if val else None,
                      fmt="₹#,##0", align="right", bg=bg)

    # Totals row
    last = len(months_sorted) + 2
    ws.row_dimensions[last].height = 18
    total_row_cols = ["Total", debits.amount.sum(),
                      credits.amount.sum(),
                      credits.amount.sum() - debits.amount.sum(),
                      len(df)]
    for ci, val in enumerate(total_row_cols, 1):
        c = ws.cell(row=last, column=ci, value=val)
        c.font   = _font(bold=True, color="FFFFFF", size=10)
        c.fill   = _fill(CLR["header_bg"])
        c.border = _border()
        if ci > 1 and ci <= 4:
            c.number_format = "₹#,##0"
        c.alignment = _align("right" if ci > 1 else "left")

    # Column widths
    ws.column_dimensions["A"].width = 14
    for i in range(2, 6 + len(pivot.columns)):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "B2"
    return ws


def build_transactions_sheet(wb, df):
    ws = wb.create_sheet("All Transactions")
    ws.sheet_view.showGridLines = False

    cols = ["Date", "Description", "Category", "Type", "Amount (₹)",
            "Source", "Month"]
    header_row(ws, 1, cols)

    for ri, (_, row) in enumerate(df.sort_values("date", ascending=False)
                                    .iterrows(), 2):
        bg = "FFFFFF" if ri % 2 == 0 else CLR["alt_row"]
        data_cell(ws, ri, 1, row.date, bg=bg)
        data_cell(ws, ri, 2, row.description, bg=bg)
        cat_color = CAT_COLORS.get(row.category, "B4B2A9")
        cc = ws.cell(row=ri, column=3, value=row.category)
        cc.font      = _font(size=10, color=cat_color, bold=True)
        cc.fill      = _fill(bg)
        cc.border    = _border("bottom", CLR["border"])
        cc.alignment = _align()
        typ = "Debit" if row.is_debit else "Credit"
        data_cell(ws, ri, 4, typ, align="center",
                  color=CLR["debit"] if row.is_debit else CLR["credit"], bg=bg)
        data_cell(ws, ri, 5, row.amount, fmt="₹#,##0", align="right",
                  color=CLR["debit"] if row.is_debit else CLR["credit"], bg=bg)
        data_cell(ws, ri, 6, row.source_name, bg=bg)
        data_cell(ws, ri, 7, row.month, bg=bg)

    set_col_widths(ws, {"A": 14, "B": 48, "C": 22, "D": 10,
                        "E": 16, "F": 22, "G": 14})
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:G1"
    return ws


def build_top10_sheet(wb, df):
    ws = wb.create_sheet("Top 10 Spends")
    ws.sheet_view.showGridLines = False

    debits = df[df.is_debit].nlargest(10, "amount").reset_index(drop=True)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Top 10 Transactions by Amount"
    t.font      = _font(bold=True, size=14, color=CLR["header_fg"])
    t.fill      = _fill(CLR["header_bg"])
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    header_row(ws, 2, ["Rank", "Date", "Description", "Category",
                        "Amount (₹)", "Source"])

    for ri, row in debits.iterrows():
        rn = ri + 3
        bg = CLR["warn_bg"] if ri == 0 else ("FFFFFF" if rn % 2 == 0 else CLR["alt_row"])
        rank_c = ws.cell(row=rn, column=1, value=f"#{ri+1}")
        rank_c.font      = _font(bold=True, size=11,
                                  color="B45309" if ri == 0 else CLR["neutral"])
        rank_c.fill      = _fill(bg)
        rank_c.alignment = _align("center")
        rank_c.border    = _border("bottom", CLR["border"])

        data_cell(ws, rn, 2, row.date, bg=bg)
        data_cell(ws, rn, 3, row.description, bg=bg)
        cat_c = ws.cell(row=rn, column=4, value=row.category)
        cat_c.font      = _font(size=10, color=CAT_COLORS.get(row.category, "B4B2A9"), bold=True)
        cat_c.fill      = _fill(bg)
        cat_c.border    = _border("bottom", CLR["border"])
        cat_c.alignment = _align()
        data_cell(ws, rn, 5, row.amount, fmt="₹#,##0", align="right",
                  color=CLR["debit"], bold=(ri == 0), bg=bg)
        data_cell(ws, rn, 6, row.source_name, bg=bg)

    set_col_widths(ws, {"A": 8, "B": 14, "C": 48, "D": 22, "E": 16, "F": 22})
    return ws


def build_extraordinary_sheet(wb, df):
    ws = wb.create_sheet("Extraordinary Transactions")
    ws.sheet_view.showGridLines = False

    debits      = df[df.is_debit]
    months      = df.month.nunique()
    avg_monthly = debits.amount.sum() / months if months else 0
    threshold   = avg_monthly * 0.30

    extraordinary = (debits[debits.amount >= threshold]
                     .sort_values("amount", ascending=False)
                     .reset_index(drop=True))

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Extraordinary Transactions"
    t.font  = _font(bold=True, size=14, color=CLR["header_fg"])
    t.fill  = _fill(CLR["header_bg"])
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"Transactions ≥ 30% of average monthly spend  "
               f"(threshold: ₹{threshold:,.0f} | avg monthly: ₹{avg_monthly:,.0f})")
    s.font      = _font(size=10, color="92400E")
    s.fill      = _fill("FEF3C7")
    s.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    header_row(ws, 3, ["Date", "Description", "Category", "Amount (₹)",
                        "Month", "% of Monthly Avg", "Source"])

    for ri, row in extraordinary.iterrows():
        rn = ri + 4
        bg = "FFFFFF" if rn % 2 == 0 else "FFF7ED"
        pct = row.amount / avg_monthly if avg_monthly else 0
        data_cell(ws, rn, 1, row.date, bg=bg)
        data_cell(ws, rn, 2, row.description, bg=bg)
        cc = ws.cell(row=rn, column=3, value=row.category)
        cc.font      = _font(size=10, color=CAT_COLORS.get(row.category, "B4B2A9"), bold=True)
        cc.fill      = _fill(bg)
        cc.border    = _border("bottom", CLR["border"])
        cc.alignment = _align()
        data_cell(ws, rn, 4, row.amount, fmt="₹#,##0", align="right",
                  color=CLR["debit"], bold=True, bg=bg)
        data_cell(ws, rn, 5, row.month, bg=bg)
        pct_c = ws.cell(row=rn, column=6, value=pct)
        pct_c.number_format = "0%"
        pct_c.font      = _font(size=10, color="DC2626" if pct > 1 else "92400E", bold=True)
        pct_c.fill      = _fill(bg)
        pct_c.border    = _border("bottom", CLR["border"])
        pct_c.alignment = _align("center")
        data_cell(ws, rn, 7, row.source_name, bg=bg)

    if len(extraordinary) == 0:
        ws.merge_cells("A4:G4")
        ws["A4"].value = "✅  No extraordinary transactions found — great spending discipline!"
        ws["A4"].font  = _font(size=11, color="166534")
        ws["A4"].fill  = _fill("DCFCE7")
        ws["A4"].alignment = _align("center")

    set_col_widths(ws, {"A": 14, "B": 48, "C": 22, "D": 16,
                        "E": 14, "F": 18, "G": 22})
    return ws



def build_dashboard_sheet(wb, df):
    """
    Pure cell-based dark dashboard — no openpyxl charts.
    Uses coloured cells, Unicode block characters, and formatted tables
    so everything renders perfectly without chart/drawing layer issues.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "4F8EF7"

    # ── Palette ────────────────────────────────────────────────────────────────
    BG       = "111827"   # page background
    PANEL    = "1F2937"   # panel / card background
    PANEL2   = "374151"   # alternate row
    SEP      = "4B5563"   # separator lines
    RED      = "F87171"
    GREEN    = "4ADE80"
    BLUE     = "60A5FA"
    AMBER    = "FBBF24"
    PURPLE   = "A78BFA"
    CYAN     = "22D3EE"
    WHITE    = "F9FAFB"
    GREY     = "9CA3AF"
    DGREY    = "6B7280"

    CAT_HEX = {
        "Food & Dining":"60A5FA","Shopping":"4ADE80","Transport":"F87171",
        "Entertainment":"A78BFA","Utilities":"FBBF24","Health":"F472B6",
        "Travel":"22D3EE","Salary/Income":"4ADE80","EMI/Loan":"FB7185",
        "Transfers":"94A3B8","Investment":"34D399","Rent":"C084FC",
        "Education":"F97316","Subscriptions":"06B6D4","Kids & Family":"FBBF24",
        "Personal Care":"EC4899","ATM & Cash":"9CA3AF","Charity":"86EFAC",
        "Other":"6B7280",
    }

    def F(h):   return PatternFill("solid", fgColor=h)
    def Ft(bold=False, sz=10, color=WHITE, italic=False):
        return Font(bold=bold, size=sz, color=color, italic=italic, name="Calibri")
    def Al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def Br(color=SEP, sides="all"):
        s = Side(style="thin", color=color)
        n = Side(style=None)
        if sides == "all":  return Border(left=s, right=s, top=s, bottom=s)
        if sides == "b":    return Border(bottom=s)
        if sides == "t":    return Border(top=s)
        return Border()

    def paint(r1, c1, r2, c2, bg=BG):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = F(bg)

    def cell(r, c, val="", bold=False, sz=10, color=WHITE, bg=PANEL,
             align="left", italic=False, fmt=None, border=None):
        cl = ws.cell(r, c, value=val)
        cl.font      = Ft(bold, sz, color, italic)
        cl.fill      = F(bg)
        cl.alignment = Al(align)
        if fmt:    cl.number_format = fmt
        if border: cl.border = border
        return cl

    def mspan(r, c1, c2, val="", bold=False, sz=10, color=WHITE,
              bg=PANEL, align="center", italic=False):
        ws.merge_cells(start_row=r, start_column=c1,
                       end_row=r,   end_column=c2)
        return cell(r, c1, val, bold, sz, color, bg, align, italic)

    # ── Column layout: 48 cols, varying widths ─────────────────────────────────
    # Left panel  (charts): cols 1-28
    # Right panel (table) : cols 29-48
    TOTAL_C  = 48
    SPLIT_C  = 29     # right panel starts here
    RIGHT_W  = TOTAL_C

    col_widths = {}
    for c in range(1, TOTAL_C+1):
        if c < SPLIT_C:
            col_widths[c] = 3.0    # left panel narrow cols for bar chart blocks
        else:
            col_widths[c] = 3.8    # right panel

    # Specific width overrides
    for c in [1,2]: col_widths[c] = 1.5   # rank/icon cols
    col_widths[SPLIT_C] = 2.0              # separator

    for c, w in col_widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w

    # Paint full canvas
    paint(1, 1, 80, TOTAL_C, BG)

    # ── Precompute ─────────────────────────────────────────────────────────────
    def month_sort(m):
        try: return __import__("datetime").datetime.strptime(m, "%b %Y")
        except: return __import__("datetime").datetime.min

    debits        = df[df.is_debit]
    credits       = df[~df.is_debit]
    months_sorted = sorted(df.month.unique(), key=month_sort)
    n_months      = max(len(months_sorted), 1)
    total_spend   = debits.amount.sum()
    total_income  = credits.amount.sum()
    net_savings   = total_income - total_spend
    avg_spend     = total_spend / n_months
    avg_income    = total_income / n_months
    sav_rate      = (net_savings / total_income * 100) if total_income else 0

    cat_totals = (debits.groupby("category")["amount"].sum()
                  .sort_values(ascending=False))
    inc_totals = (credits.groupby("category")["amount"].sum()
                  .sort_values(ascending=False))
    month_sp   = debits.groupby("month")["amount"].sum()
    month_in   = credits.groupby("month")["amount"].sum()

    top10 = (debits.groupby(["description","category"], as_index=False)["amount"]
             .max().nlargest(10, "amount").reset_index(drop=True))

    # ════════════════════════════════════════════════════════════════════════════
    # ROW 1-2 : TITLE BANNER
    # ════════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 36
    paint(1, 1, 2, TOTAL_C, "0F172A")

    title_txt = (f"  PERSONAL FINANCE DASHBOARD"
                 f"          {datetime.now():%d %b %Y}")
    mspan(2, 1, TOTAL_C, title_txt, bold=True, sz=16,
          color=WHITE, bg="0F172A", align="left")

    # ════════════════════════════════════════════════════════════════════════════
    # ROW 3-5 : KPI CARDS  (6 cards)
    # ════════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 4
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28
    paint(3, 1, 5, TOTAL_C, BG)

    kpis = [
        ("TOTAL INCOME",      f"\u20b9{total_income:,.0f}",  GREEN,  "1F3B2D"),
        ("TOTAL SPEND",       f"\u20b9{total_spend:,.0f}",   RED,    "3B1F1F"),
        ("NET SAVINGS",       f"\u20b9{abs(net_savings):,.0f}",
                              GREEN if net_savings>=0 else RED,
                              "1F3B2D" if net_savings>=0 else "3B1F1F"),
        ("AVG MONTHLY SPEND", f"\u20b9{avg_spend:,.0f}",     AMBER,  "3B2F1A"),
        ("SAVINGS RATE",      f"{sav_rate:.1f}%",             BLUE,   "1A2340"),
        ("TRANSACTIONS",      f"{len(df):,}",                 PURPLE, "2A1F40"),
    ]
    kpi_w = TOTAL_C // len(kpis)
    for ki, (lbl, val, col, bg_dark) in enumerate(kpis):
        c1 = ki * kpi_w + 1
        c2 = c1 + kpi_w - 1
        if ki == len(kpis)-1: c2 = TOTAL_C
        # top accent line
        paint(3, c1, 3, c2, col)
        # label row
        paint(4, c1, 4, c2, bg_dark)
        mspan(4, c1, c2, lbl, bold=False, sz=8, color=GREY, bg=bg_dark)
        # value row
        paint(5, c1, 5, c2, bg_dark)
        mspan(5, c1, c2, val, bold=True, sz=14, color=col, bg=bg_dark)

    ws.row_dimensions[6].height = 6
    paint(6, 1, 6, TOTAL_C, BG)

    # ════════════════════════════════════════════════════════════════════════════
    # LEFT PANEL — rows 7 onward, cols 1 to SPLIT_C-1
    # ════════════════════════════════════════════════════════════════════════════
    L1 = 1
    L2 = SPLIT_C - 2    # last col of left panel content
    BAR_END  = L2        # bar chart ends here
    BAR_COLS = BAR_END - 6   # number of cols available for bar fill

    # ── SECTION A: Spend by Category (horizontal cell bar chart) ──────────────
    ws.row_dimensions[7].height = 18
    paint(7, L1, 7, L2, PANEL)
    mspan(7, L1, L2, "  SPEND BY CATEGORY", bold=True, sz=10,
          color=AMBER, bg=PANEL, align="left")

    ws.row_dimensions[8].height = 13
    paint(8, L1, 8, L2, PANEL2)
    for c, hdr in zip([3, 5, BAR_END-5, BAR_END-2],
                      ["Category", "", "Amount", "%"]):
        cell(8, c, hdr, bold=True, sz=8, color=GREY, bg=PANEL2, align="left")

    max_cat = cat_totals.iloc[0] if not cat_totals.empty else 1
    total_spend_nz = total_spend or 1

    for ri, (cat, amt) in enumerate(cat_totals.head(12).items()):
        rn = 9 + ri
        ws.row_dimensions[rn].height = 15
        bg_r = PANEL if ri % 2 == 0 else PANEL2
        paint(rn, L1, rn, L2, bg_r)

        hex_c = CAT_HEX.get(cat, GREY)
        # colour swatch col 1
        ws.cell(rn, 1).fill = F(hex_c)

        # category name cols 2-5
        mspan(rn, 2, 5, f" {cat[:16]}", bold=False, sz=9,
              color=hex_c, bg=bg_r, align="left")

        # bar: cols 6 to BAR_END-6
        bar_len = int(amt / max_cat * (BAR_COLS - 6))
        bar_len = max(bar_len, 1)
        for bc in range(6, 6 + bar_len):
            if bc <= BAR_END - 6:
                ws.cell(rn, bc).fill = F(hex_c)
        # empty rest of bar
        for bc in range(6 + bar_len, BAR_END - 5):
            ws.cell(rn, bc).fill = F(bg_r)

        # amount
        mspan(rn, BAR_END-5, BAR_END-3,
              f"\u20b9{amt:,.0f}", bold=True, sz=9,
              color=WHITE, bg=bg_r, align="right")
        # pct
        mspan(rn, BAR_END-2, BAR_END,
              f"{amt/total_spend_nz*100:.1f}%", bold=False, sz=9,
              color=GREY, bg=bg_r, align="right")

    row_after_cats = 9 + min(len(cat_totals), 12)

    # ── SECTION B: Monthly Spend vs Income ────────────────────────────────────
    ws.row_dimensions[row_after_cats].height = 8
    paint(row_after_cats, L1, row_after_cats, L2, BG)

    sh_row = row_after_cats + 1
    ws.row_dimensions[sh_row].height = 18
    paint(sh_row, L1, sh_row, L2, PANEL)
    mspan(sh_row, L1, L2, "  MONTHLY SPEND vs INCOME",
          bold=True, sz=10, color=CYAN, bg=PANEL, align="left")

    max_monthly = max((max(month_sp.max() if not month_sp.empty else 0,
                           month_in.max() if not month_in.empty else 0)), 1)
    MONTH_BAR_COLS = L2 - 10

    for mi, month in enumerate(months_sorted):
        sp_val = float(month_sp.get(month, 0))
        in_val = float(month_in.get(month, 0))
        net_v  = in_val - sp_val

        base_r = sh_row + 1 + mi * 3
        ws.row_dimensions[base_r].height   = 10   # spend bar
        ws.row_dimensions[base_r+1].height = 10   # income bar
        ws.row_dimensions[base_r+2].height = 6    # gap

        paint(base_r,   L1, base_r,   L2, PANEL)
        paint(base_r+1, L1, base_r+1, L2, PANEL)
        paint(base_r+2, L1, base_r+2, L2, BG)

        # month label
        mspan(base_r, 1, 4, month[:7], bold=True, sz=8,
              color=WHITE, bg=PANEL, align="left")
        mspan(base_r+1, 1, 4, "", bg=PANEL)

        # spend bar (red)
        sp_len = max(int(sp_val / max_monthly * MONTH_BAR_COLS), 1)
        for bc in range(5, 5 + sp_len):
            if bc <= L2 - 6: ws.cell(base_r, bc).fill = F(RED)
        for bc in range(5 + sp_len, L2 - 5):
            ws.cell(base_r, bc).fill = F(PANEL)
        mspan(base_r, L2-5, L2-1,
              f"\u20b9{sp_val:,.0f}", sz=8, color=RED,
              bg=PANEL, align="right")

        # income bar (green)
        in_len = max(int(in_val / max_monthly * MONTH_BAR_COLS), 1)
        for bc in range(5, 5 + in_len):
            if bc <= L2 - 6: ws.cell(base_r+1, bc).fill = F(GREEN)
        for bc in range(5 + in_len, L2 - 5):
            ws.cell(base_r+1, bc).fill = F(PANEL)
        mspan(base_r+1, L2-5, L2-1,
              f"\u20b9{in_val:,.0f}", sz=8, color=GREEN,
              bg=PANEL, align="right")

    # ── SECTION C: Net savings sparkline ──────────────────────────────────────
    last_month_row = sh_row + 1 + len(months_sorted) * 3
    ws.row_dimensions[last_month_row].height = 8
    paint(last_month_row, L1, last_month_row, L2, BG)

    sp2 = last_month_row + 1
    ws.row_dimensions[sp2].height = 18
    paint(sp2, L1, sp2, L2, PANEL)
    mspan(sp2, L1, L2, "  NET SAVINGS TREND",
          bold=True, sz=10, color=PURPLE, bg=PANEL, align="left")

    net_vals = [float(month_in.get(m, 0)) - float(month_sp.get(m, 0))
                for m in months_sorted]
    if net_vals:
        nv_min = min(net_vals)
        nv_max = max(net_vals)
        nv_range = (nv_max - nv_min) or 1
        SPARK_ROWS = 8   # height in rows for sparkline area
        for sr in range(sp2+1, sp2+1+SPARK_ROWS):
            ws.row_dimensions[sr].height = 10
            paint(sr, L1, sr, L2, PANEL)

        n_m = len(months_sorted)
        col_per = max(1, (L2 - 2) // n_m) if n_m else 1
        for mi, (month, nv) in enumerate(zip(months_sorted, net_vals)):
            c_center = 2 + mi * col_per
            # normalise to 0-SPARK_ROWS
            norm = (nv - nv_min) / nv_range
            filled = max(1, round(norm * (SPARK_ROWS - 1)))
            dot_row = sp2 + SPARK_ROWS - filled  # higher value = higher up
            col_c = CAT_HEX.get("", "")
            dot_color = GREEN if nv >= 0 else RED
            for cc in range(c_center, min(c_center + col_per - 1, L2)):
                ws.cell(dot_row, cc).fill = F(dot_color)
            # month label at bottom
            br = sp2 + SPARK_ROWS
            ws.row_dimensions[br].height = 12
            mspan(br, c_center, min(c_center+col_per-1, L2),
                  month[:3], sz=7, color=GREY, bg=PANEL, align="center")
            # net value label
            mspan(dot_row, c_center, min(c_center+col_per-1, L2),
                  f"\u20b9{nv/1000:.0f}k" if abs(nv) >= 1000 else f"\u20b9{nv:.0f}",
                  sz=7, color=dot_color, bg=PANEL, align="center")

    # ════════════════════════════════════════════════════════════════════════════
    # SEPARATOR COLUMN
    # ════════════════════════════════════════════════════════════════════════════
    for r in range(7, 80):
        ws.cell(r, SPLIT_C-1).fill = F(SEP)
        ws.row_dimensions[r].height = ws.row_dimensions[r].height or 15

    # ════════════════════════════════════════════════════════════════════════════
    # RIGHT PANEL — TOP 10 SPENDS + INCOME SUMMARY
    # ════════════════════════════════════════════════════════════════════════════
    RC = SPLIT_C + 1   # right content start col

    # header
    ws.row_dimensions[7].height = 18
    paint(7, SPLIT_C, 7, TOTAL_C, PANEL)
    mspan(7, RC, TOTAL_C, "  TOP 10 SPENDS",
          bold=True, sz=10, color=AMBER, bg=PANEL, align="left")

    # column headers
    ws.row_dimensions[8].height = 13
    paint(8, SPLIT_C, 8, TOTAL_C, PANEL2)
    mspan(8, RC,         RC+1,        "#",    bold=True, sz=8, color=GREY, bg=PANEL2, align="center")
    mspan(8, RC+2,       TOTAL_C-5,  "Description", bold=True, sz=8, color=GREY, bg=PANEL2)
    mspan(8, TOTAL_C-4, TOTAL_C,    "Amount", bold=True, sz=8, color=GREY, bg=PANEL2, align="right")

    for ti, trow in top10.iterrows():
        rn = 9 + ti * 2
        ws.row_dimensions[rn].height   = 16
        ws.row_dimensions[rn+1].height = 11
        bg_r = PANEL if ti % 2 == 0 else PANEL2
        paint(rn,   SPLIT_C, rn,   TOTAL_C, bg_r)
        paint(rn+1, SPLIT_C, rn+1, TOTAL_C, bg_r)

        hex_c    = CAT_HEX.get(trow.category, GREY)
        raw_desc = str(trow.description)
        desc_s   = (raw_desc[:22] + "\u2026") if len(raw_desc) > 23 else raw_desc

        # rank badge
        mspan(rn, RC, RC+1, f"#{ti+1}",
              bold=True, sz=9, color=AMBER if ti==0 else GREY,
              bg=bg_r, align="center")
        # description
        mspan(rn, RC+2, TOTAL_C-5, f" {desc_s}",
              bold=False, sz=9, color=WHITE, bg=bg_r)
        # amount
        mspan(rn, TOTAL_C-4, TOTAL_C,
              f"\u20b9{trow.amount:,.0f}",
              bold=True, sz=10, color=RED, bg=bg_r, align="right")
        # category
        mspan(rn+1, RC, TOTAL_C, f"   {trow.category}",
              bold=False, sz=7, color=hex_c, bg=bg_r, italic=True)

        # separator
        for sc in range(SPLIT_C, TOTAL_C+1):
            ws.cell(rn+2, sc).fill = F(SEP) if sc > SPLIT_C else F(BG)
        ws.row_dimensions[rn+2].height = 2

    # ── INCOME SUMMARY (below top-10) ─────────────────────────────────────────
    inc_start = 9 + 10*3 + 1
    ws.row_dimensions[inc_start].height = 8
    paint(inc_start, SPLIT_C, inc_start, TOTAL_C, BG)

    ws.row_dimensions[inc_start+1].height = 18
    paint(inc_start+1, SPLIT_C, inc_start+1, TOTAL_C, PANEL)
    mspan(inc_start+1, RC, TOTAL_C, "  INCOME SOURCES",
          bold=True, sz=10, color=GREEN, bg=PANEL, align="left")

    max_inc = inc_totals.iloc[0] if not inc_totals.empty else 1
    INC_BAR = TOTAL_C - RC - 8

    for ii, (cat, amt) in enumerate(inc_totals.head(8).items()):
        rn = inc_start + 2 + ii * 2
        ws.row_dimensions[rn].height   = 14
        ws.row_dimensions[rn+1].height = 5
        bg_r = PANEL if ii % 2 == 0 else PANEL2
        paint(rn,   SPLIT_C, rn,   TOTAL_C, bg_r)
        paint(rn+1, SPLIT_C, rn+1, TOTAL_C, bg_r)

        hex_c   = CAT_HEX.get(cat, GREEN)
        bar_len = max(int(amt / max_inc * INC_BAR), 1)

        # swatch + label
        ws.cell(rn, RC).fill = F(hex_c)
        mspan(rn, RC+1, RC+5, f" {cat[:14]}",
              sz=8, color=hex_c, bg=bg_r)
        # bar
        for bc in range(RC+6, RC+6+bar_len):
            if bc <= TOTAL_C-5:
                ws.cell(rn, bc).fill = F(hex_c)
        for bc in range(RC+6+bar_len, TOTAL_C-4):
            ws.cell(rn, bc).fill = F(bg_r)
        # amount
        mspan(rn, TOTAL_C-4, TOTAL_C,
              f"\u20b9{amt:,.0f}", bold=True, sz=9,
              color=GREEN, bg=bg_r, align="right")

    ws.freeze_panes = "A7"
    return ws


def build_excel(df, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, df)
    build_monthly_sheet(wb, df)
    build_transactions_sheet(wb, df)
    build_top10_sheet(wb, df)
    build_extraordinary_sheet(wb, df)
    build_dashboard_sheet(wb, df)

    wb.save(output_path)
    print(f"\n✅  Dashboard saved → {output_path}")

# ═══════════════════════════════════════════════════════════════════════════════
#  5. INTERACTIVE CLI
# ═══════════════════════════════════════════════════════════════════════════════

def prompt(msg, default=None):
    suffix = f" [{default}]" if default else ""
    val = input(f"  {msg}{suffix}: ").strip()
    return val if val else default

def configure_source_interactively(configs):
    """Walk user through configuring a new/existing statement type."""
    print("\n─── Statement Configuration ────────────────────────────────")
    print("  Known types:", ", ".join(configs.keys()))
    name = prompt("Choose existing name OR enter new name", "Custom")

    cfg = dict(configs.get(name, configs.get("Custom", {})))
    cfg["name"] = name

    print(f"\n  Configuring: {name}")
    cfg["type"] = prompt("Type (bank / credit)", cfg.get("type", "bank"))

    print("\n  Column mapping (enter column name as it appears in the CSV header,")
    print("  or enter column NUMBER like '1', '2', '3' ...)")
    cfg["date_col"]    = prompt("Date column", cfg.get("date_col", "Date"))
    cfg["date_format"] = prompt("Date format (e.g. %d/%m/%Y  %d-%m-%Y  %Y-%m-%d)",
                                cfg.get("date_format", "%d/%m/%Y"))
    cfg["desc_col"]    = prompt("Description/Narration column", cfg.get("desc_col", ""))
    cfg["skip_rows"]   = int(prompt("Rows to skip before header", str(cfg.get("skip_rows", 0))))

    fmt = prompt("Amount format  1=Separate Debit/Credit cols  2=Single signed col",
                 "1" if cfg.get("amount_format") != "signed" else "2")
    if fmt == "2":
        cfg["amount_format"] = "signed"
        cfg["amount_col"]    = prompt("Amount column", cfg.get("amount_col", "Amount"))
        sign = prompt("Sign convention  1=Positive means Debit  2=Negative means Debit", "1")
        cfg["sign_convention"] = "pos_is_debit" if sign == "1" else "neg_is_debit"
    else:
        cfg["amount_format"] = "separate"
        cfg["debit_col"]  = prompt("Debit/Withdrawal column", cfg.get("debit_col", ""))
        cfg["credit_col"] = prompt("Credit/Deposit column",  cfg.get("credit_col", ""))

    configs[name] = cfg
    return name, cfg


def interactive_mode():
    print("\u2554" + "\u2550"*58 + "\u2557")
    print("\u2551         Bank Statement Analyzer  \u2014  Local Tool           \u2551")
    print("\u255a" + "\u2550"*58 + "\u255d")
    print(f"  Keyword mapping : {KEYWORDS_CSV}")
    print(f"  Configs file    : {CONFIGS_FILE}")
    print(f"  Session file    : {SESSION_FILE}")

    keyword_map = load_keywords()
    configs     = load_configs()
    session     = load_session()

    # ── Restore previously saved files ───────────────────────────────────────
    file_cfgs     = []   # list of (Path, cfg_dict)
    restored_keys = set()

    if session["files"]:
        existing = []
        missing  = []
        for e in session["files"]:
            p = Path(e["path"])
            if p.exists():
                existing.append((str(p), e["cfg"]))
            else:
                missing.append(str(p))

        print(f"\n  {'─'*56}")
        print(f"  Last session: {len(session['files'])} file(s)")
        print(f"  {'─'*56}")
        for i, e in enumerate(session["files"], 1):
            p = Path(e["path"])
            status = "\u2713" if p.exists() else "\u2717 NOT FOUND"
            print(f"    {i:>2}.  {status}  {p.name:<38} [{e['cfg'].get('name','')}]")
        print(f"  {'─'*56}")

        if missing:
            print(f"  Note: {len(missing)} file(s) marked NOT FOUND may have been moved/renamed.")

        if existing:
            print()
            print("  Options:")
            print("    a  = load ALL found files  (default)")
            print("    n  = load NONE, start fresh")
            print("    1,3,5 ... = load only the files with those numbers")
            choice = prompt("  Your choice", "a").strip().lower()

            if choice == "a":
                file_cfgs = [(Path(fp), cfg) for fp, cfg in existing]
                restored_keys = {session_file_key(fp) for fp, _ in file_cfgs}
                print(f"  Loaded all {len(file_cfgs)} file(s).")
            elif choice == "n":
                print("  Starting fresh — no files loaded from session.")
            else:
                # Parse comma-separated numbers
                try:
                    chosen_nums = {int(x.strip()) for x in choice.split(",") if x.strip()}
                    # Map 1-based index to existing[] (only found files are candidates)
                    # Use the original session order for numbering
                    chosen_files = []
                    found_set = {str(Path(fp).resolve()) for fp, _ in existing}
                    ei = 0  # index into existing[]
                    for idx, e in enumerate(session["files"], 1):
                        p = Path(e["path"])
                        if p.exists():
                            if idx in chosen_nums:
                                chosen_files.append((str(p), e["cfg"]))
                            ei += 1
                    if chosen_files:
                        file_cfgs = [(Path(fp), cfg) for fp, cfg in chosen_files]
                        restored_keys = {session_file_key(fp) for fp, _ in file_cfgs}
                        print(f"  Loaded {len(file_cfgs)} selected file(s):")
                        for fp, _ in file_cfgs:
                            print(f"    \u2713  {Path(fp).name}")
                    else:
                        print("  No valid numbers matched — starting fresh.")
                except ValueError:
                    print("  Could not parse input — loading all found files.")
                    file_cfgs = [(Path(fp), cfg) for fp, cfg in existing]
                    restored_keys = {session_file_key(fp) for fp, _ in file_cfgs}

            # Warn if keywords CSV changed since last run
            if file_cfgs:
                saved_mtime = session.get("keywords_mtime")
                cur_mtime = os.path.getmtime(KEYWORDS_CSV) if KEYWORDS_CSV.exists() else None
                if saved_mtime and cur_mtime and abs(cur_mtime - saved_mtime) > 1:
                    print("\n  [INFO] category_keywords.csv was modified since last run.")
                    print("         Transactions will be re-categorized with updated keywords.")
        else:
            print("  None of the saved files exist on disk — starting fresh.")

    # ── Let user add new files ────────────────────────────────────────────────
    print("\n  Add new statement files below. Press ENTER with no path when done.")
    already_added = {session_file_key(fp) for fp, _ in file_cfgs}

    while True:
        print("\n\u250c\u2500 Add statement file " + "\u2500"*38)
        filepath_raw = prompt("Path to CSV file (or ENTER to finish)", "")
        if not filepath_raw:
            if not file_cfgs:
                print("  No files added. Exiting.")
                return
            break

        filepath = Path(filepath_raw.strip('"').strip("'"))
        if not filepath.exists():
            print(f"  [ERROR] File not found: {filepath}")
            continue

        key = session_file_key(filepath)
        if key in already_added:
            print(f"  [SKIP] {filepath.name} is already in the list.")
            continue

        # Peek at headers
        _peek = None
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                _peek = pd.read_csv(filepath, nrows=3, dtype=str,
                                    encoding=enc, on_bad_lines="skip")
                _peek.columns = [str(c).strip() for c in _peek.columns]
                break
            except Exception:
                continue

        print(f"\n  File: {filepath.name}")
        if _peek is not None:
            print(f"  Columns detected : {list(_peek.columns)}")
            if not _peek.empty:
                print(f"  First data row   : {_peek.iloc[0].tolist()}")

        print(f"\n  Known statement types: {', '.join(configs.keys())}")
        name = prompt("Which statement type? (or NEW to configure manually)", "HDFC Bank")

        if name.upper() == "NEW" or name not in configs:
            name, cfg = configure_source_interactively(configs)
            save_configs(configs)
        else:
            cfg = dict(configs[name])
            cfg["name"] = name
            if _peek is not None:
                _resolved = auto_detect_columns(_peek, cfg)
                missing_cols = [k for k, v in _resolved.items()
                                if k in ("date_col", "desc_col") and not v]
                if missing_cols:
                    print(f"\n  [WARN] Columns for {missing_cols} not found using '{name}' preset.")
                    fix = prompt("  Reconfigure column mapping now? (y/n)", "y")
                    if fix.lower() == "y":
                        name, cfg = configure_source_interactively(configs)
                        save_configs(configs)
                else:
                    d = _resolved.get("date_col", "?")
                    s = _resolved.get("desc_col", "?")
                    print(f"  Auto-matched : date='{d}'  desc='{s}'")

        file_cfgs.append((filepath, cfg))
        already_added.add(key)
        print(f"  \u2713 Added {filepath.name} as '{name}'")

    # ── Show final file list before parsing ──────────────────────────────────
    if file_cfgs:
        new_files = [(fp, cfg) for fp, cfg in file_cfgs
                     if session_file_key(fp) not in restored_keys]
        if new_files:
            print(f"\n  Final list: {len(file_cfgs)} file(s) "
                  f"({len(file_cfgs)-len(new_files)} restored + {len(new_files)} new)")
        else:
            print(f"\n  Final list: {len(file_cfgs)} file(s) (all restored from session)")
        for i, (fp, cfg) in enumerate(file_cfgs, 1):
            tag = "[new]" if session_file_key(fp) not in restored_keys else ""
            print(f"    {i:>2}.  {Path(fp).name:<42} [{cfg.get('name','')}]  {tag}")

    # ── Parse all files ───────────────────────────────────────────────────────
    print("\n\u23f3  Parsing files \u2026")
    all_txns = []
    for filepath, cfg in file_cfgs:
        rows = parse_statement(filepath, cfg, keyword_map)
        all_txns.extend(rows)
        print(f"  \u2713 {Path(filepath).name:<45} -> {len(rows):>4} transactions")

    if not all_txns:
        print("\n[ERROR] No transactions parsed. Check column names and date format.")
        return

    df = pd.DataFrame(all_txns)
    debits  = df[df.is_debit]
    credits = df[~df.is_debit]

    # Quick preview
    print(f"\n{'\u2500'*56}")
    print(f"  Total transactions : {len(df):,}")
    print(f"  Total spend        : Rs.{debits.amount.sum():,.0f}")
    print(f"  Total income       : Rs.{credits.amount.sum():,.0f}")
    print(f"  Net savings        : Rs.{credits.amount.sum()-debits.amount.sum():,.0f}")
    print(f"  Date range         : {df.date.min()}  to  {df.date.max()}")
    print(f"\n  Category breakdown (debits):")
    cat_grp = debits.groupby("category")["amount"].sum().sort_values(ascending=False)
    for cat, amt in cat_grp.items():
        bar = "\u2588" * min(int(amt / cat_grp.max() * 20), 20)
        print(f"    {cat:<22} {bar:<20} Rs.{amt:>10,.0f}")

    # Output path — default to last used path
    last_out = session.get("output", "")
    out_default = last_out if last_out else str(
        Path.cwd() / f"bank_dashboard_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    out_path = prompt(f"\nOutput Excel file path", out_default)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    build_excel(df, out_path)

    # ── Save session so next run remembers all files ──────────────────────────
    save_session(file_cfgs, out_path)

    print(f"\n  Sheets: Summary | Monthly Breakdown | All Transactions | Top 10 Spends | Extraordinary | Charts")
    print(f"  Next time you run, all {len(file_cfgs)} file(s) will be pre-loaded automatically.")




# ═══════════════════════════════════════════════════════════════════════════════
#  6. ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Bank Statement Analyzer")
    parser.add_argument("--config", help="JSON config file with file→source mappings")
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            run_cfg = json.load(f)
        keyword_map = load_keywords()
        configs     = load_configs()
        all_txns    = []
        for entry in run_cfg.get("files", []):
            cfg = dict(configs.get(entry["source"], configs["Custom"]))
            cfg["name"] = entry["source"]
            rows = parse_statement(entry["file"], cfg, keyword_map)
            all_txns.extend(rows)
        df = pd.DataFrame(all_txns)
        out = run_cfg.get("output", f"bank_dashboard_{datetime.now():%Y%m%d}.xlsx")
        build_excel(df, Path(out))
    else:
        try:
            interactive_mode()
        except (KeyboardInterrupt, EOFError):
            print("\n\nBye!")

if __name__ == "__main__":
    main()
